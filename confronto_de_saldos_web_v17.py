# confronto_de_saldos_web_v17.py
# Streamlit web app com OCR automático + AgGrid + export Excel/PDF
# Uso: streamlit run confronto_de_saldos_web_v17.py

import io
import re
import os
import sys
import tempfile
import base64
import threading
from typing import List, Optional, Tuple
import streamlit as st
import pandas as pd

# libs para PDF/text/OCR/export
try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    from pdf2image import convert_from_bytes
except Exception:
    convert_from_bytes = None

try:
    import pytesseract
except Exception:
    pytesseract = None

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
except Exception:
    pass

# AgGrid
try:
    from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
    from st_aggrid.shared import GridUpdateMode
except Exception:
    AgGrid = None
    GridOptionsBuilder = None
    JsCode = None
    GridUpdateMode = None

# -----------------------
# Configs OCR: ajustável
# -----------------------
# Se Tesseract não estiver no PATH, descomente e ajuste abaixo
# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# -----------------------
# Utilitários (mesma lógica)
# -----------------------
CURRENCY_RE = re.compile(r'(-?\d{1,3}(?:[.\s]\d{3})*(?:,\d{2})|\d+(?:,\d{2}))')
LABELS_PREV = [r'SALDO ANTERIOR', r'Saldo anterior', r'Saldo Anterior']
LABELS_CURR = [r'SALDO ATUAL', r'Saldo atual', r'Saldo Atual']


def parse_currency_to_float(s: str):
    if s is None:
        return None
    s = str(s)
    s = re.sub(r'[Rr]\$\s*', '', s).strip()
    m = CURRENCY_RE.search(s)
    if not m:
        s2 = re.sub(r'[^\d\.\-]', '', s)
        try:
            return float(s2) if s2 else None
        except:
            return None
    token = m.group(0)
    token = token.replace('.', '').replace(' ', '').replace(',', '.')
    try:
        return float(token)
    except:
        return None


def format_brazilian(value):
    # Esta verificação agora funciona corretamente com None (ao invés de np.nan)
    if value is None:
        return ''
    try:
        if isinstance(value, (int, float)):
            return f"{value:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")
        else:
            v = float(str(value).replace('.', '').replace(',', '.'))
            return f"{v:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")
    except Exception:
        return str(value)


def extract_text_with_pdfplumber_bytes(b: bytes) -> str:
    if not pdfplumber:
        return ""
    try:
        with pdfplumber.open(io.BytesIO(b)) as pdf:
            pages = []
            for p in pdf.pages:
                pages.append(p.extract_text() or '')
            return "\n".join(pages)
    except Exception:
        return ""


def extract_text_with_ocr_bytes(b: bytes) -> str:
    """Converte PDF (bytes) para imagens e executa pytesseract OCR em cada página."""
    if convert_from_bytes is None or pytesseract is None:
        return ""
    try:
        images = convert_from_bytes(b)
    except Exception:
        return ""
    texts = []
    for img in images:
        try:
            txt = pytesseract.image_to_string(img, lang='eng+por')  # tentar Português+Inglês
        except Exception:
            txt = ""
        texts.append(txt)
    return "\n".join(texts)


def extract_text_from_pdf_bytes_smart(b: bytes) -> str:
    """Tenta pdfplumber; se vazio, usa OCR fallback."""
    text = extract_text_with_pdfplumber_bytes(b)
    if text and len(text.strip()) > 20:
        return text
    # if no text found, try OCR
    if pytesseract is not None and convert_from_bytes is not None:
        return extract_text_with_ocr_bytes(b)
    return text or ""


def find_label_value_in_window(window: str, labels):
    for lab in labels:
        m = re.search(r'(' + re.escape(lab) + r')[^\n\r]{0,100}([^\n\r]*)', window)
        if m:
            tail = m.group(2)
            val = parse_currency_to_float(tail)
            if val is not None:
                return val
    return None


def find_balance_by_label_in_text(text: str, account: str, label_group='current'):
    if not text:
        return None
    txt = text
    acct_esc = re.escape(account.strip())
    labels = LABELS_CURR if label_group == 'current' else LABELS_PREV

    for m in re.finditer(acct_esc, txt):
        start = max(0, m.start() - 800)
        end = min(len(txt), m.end() + 800)
        window = txt[start:end]
        val = find_label_value_in_window(window, labels)
        if val is not None:
            return val

    val_global = find_label_value_in_window(txt, labels)
    if val_global is not None:
        return val_global

    for lab in labels:
        m2 = re.search(re.escape(lab) + r'\s*[=:]?\s*([Rr]?\$?[^\n\r]{0,40})', txt)
        if m2:
            val = parse_currency_to_float(m2.group(1))
            if val is not None:
                return val
    return None


# -----------------------
# Processing: cooperative cancel + progress updates
# -----------------------
def process_confronto_streamlit(df_accounts: pd.DataFrame, pdf_files_bytes: List[bytes]):
    """Processa e atualiza st.session_state['progress_percent'] e ['processed_count'] ao longo do processamento."""
    results = []
    total_prev_excel = total_prev_pdf = total_curr_excel = total_curr_pdf = 0.0
    total_accounts = len(df_accounts.index)
    processed = 0

    for idx, row in enumerate(df_accounts.itertuples(index=False), start=1):
        # cancel check
        if st.session_state.get('cancel_requested', False):
            return results, (total_prev_excel, total_prev_pdf, total_curr_excel, total_curr_pdf), True

        account = str(getattr(row, 'account')).strip()
        excel_prev = getattr(row, 'excel_prev')
        excel_curr = getattr(row, 'excel_curr')

        if excel_prev is not None:
            total_prev_excel += excel_prev
        if excel_curr is not None:
            total_curr_excel += excel_curr

        found_pdf = None
        pdf_prev = None
        pdf_curr = None

        for b in pdf_files_bytes:
            if st.session_state.get('cancel_requested', False):
                return results, (total_prev_excel, total_prev_pdf, total_curr_excel, total_curr_pdf), True
            text = extract_text_from_pdf_bytes_smart(b)
            if not text:
                continue
            if account in text:
                pdf_prev = find_balance_by_label_in_text(text, account, label_group='previous')
                pdf_curr = find_balance_by_label_in_text(text, account, label_group='current')
                found_pdf = "<uploaded_pdf>"
                if pdf_prev is not None and pdf_curr is not None:
                    break

        if pdf_prev is not None:
            total_prev_pdf += pdf_prev
        if pdf_curr is not None:
            total_curr_pdf += pdf_curr

        def cmp(a, b):
            if a is None or b is None:
                return None, 'unk'
            diff = round(a - b, 2)
            return diff, 'confere' if abs(diff) <= 0.01 else 'nao_confere'

        diff_prev, tag_prev = cmp(excel_prev, pdf_prev)
        diff_curr, tag_curr = cmp(excel_curr, pdf_curr)

        res = dict(
            account=account,
            excel_prev=excel_prev,
            pdf_prev=pdf_prev,
            diff_prev=diff_prev,
            status_prev='CONFERE' if tag_prev == 'confere' else ('NÃO CONFERE' if tag_prev == 'nao_confere' else 'NÃO ENCONTRADO'),
            excel_curr=excel_curr,
            pdf_curr=pdf_curr,
            diff_curr=diff_curr,
            status_curr='CONFERE' if tag_curr == 'confere' else ('NÃO CONFERE' if tag_curr == 'nao_confere' else 'NÃO ENCONTRADO'),
            pdf_file=''
        )
        results.append(res)

        processed += 1
        percent = int((processed / total_accounts) * 100)
        st.session_state['progress_percent'] = percent
        st.session_state['processed_count'] = processed

    st.session_state['progress_percent'] = 100
    st.session_state['processed_count'] = processed
    return results, (total_prev_excel, total_prev_pdf, total_curr_excel, total_curr_pdf), False


# -----------------------
# Export helpers (Excel/PDF)
# -----------------------
def create_excel_bytes(results: List[dict], totals: Tuple[float, float, float, float]) -> bytes:
    df = pd.DataFrame(results)
    
    # <-- CORREÇÃO: Aplicar a conversão None -> '' aqui, ANTES de formatar.
    # Isso garante que a função format_brazilian receba None e retorne '' corretamente.
    df_excel_safe = df.astype(object).where(pd.notnull(df), None)
    
    headers_map = {
        "account": "Conta",
        "excel_prev": "Excel Prev",
        "pdf_prev": "PDF Prev",
        "diff_prev": "Dif Prev",
        "status_prev": "Status Prev",
        "excel_curr": "Excel Curr",
        "pdf_curr": "PDF Curr",
        "diff_curr": "Dif Curr",
        "status_curr": "Status Curr",
        "pdf_file": "Arquivo PDF"
    }
    df_excel_safe.rename(columns=headers_map, inplace=True)

    for col in ["Excel Prev", "PDF Prev", "Dif Prev", "Excel Curr", "PDF Curr", "Dif Curr"]:
        if col in df_excel_safe.columns:
            # Usar format_brazilian, que já trata None corretamente
            df_excel_safe[col] = df_excel_safe[col].apply(format_brazilian)

    total_prev_excel, total_prev_pdf, total_curr_excel, total_curr_pdf = totals
    totals_row = {
        "Conta": "Totais",
        "Excel Prev": format_brazilian(total_prev_excel),
        "PDF Prev": format_brazilian(total_prev_pdf),
        "Dif Prev": "",
        "Status Prev": "",
        "Excel Curr": format_brazilian(total_curr_excel),
        "PDF Curr": format_brazilian(total_curr_pdf),
        "Dif Curr": "",
        "Status Curr": "",
        "Arquivo PDF": ""
    }
    df_out = pd.concat([df_excel_safe, pd.DataFrame([totals_row])], ignore_index=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Confronto de Saldos")
        ws = writer.sheets["Confronto de Saldos"]

        from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
        from openpyxl.utils import get_column_letter

        thin = Side(border_style="thin", color="000000")
        border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

        fill_confere = PatternFill(start_color="D4F7D4", end_color="D4F7D4", fill_type="solid")
        fill_nao_confere = PatternFill(start_color="F7D4D4", end_color="F7D4D4", fill_type="solid")
        fill_nao_encontrado = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
        fill_header = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        fill_totals = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

        # header
        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.alignment = align_center

        max_row = ws.max_row
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=ws.max_column):
            is_totals = (row[0].value == "Totais")
            status_prev = (row[4].value or "").strip() if row[4].value is not None else ""
            status_curr = (row[8].value or "").strip() if row[8].value is not None else ""

            if is_totals:
                fill = fill_totals
            else:
                if "NÃO CONFERE" in status_prev or "NÃO CONFERE" in status_curr:
                    fill = fill_nao_confere
                elif "CONFERE" in status_prev or "CONFERE" in status_curr:
                    fill = fill_confere
                else:
                    fill = fill_nao_encontrado

            for cell in row:
                cell.fill = fill
                cell.border = border_all
                cell.alignment = align_center
                if is_totals:
                    cell.font = Font(bold=True)

        # autofit columns
        for i, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(i)
            for cell in col_cells:
                try:
                    val = cell.value
                    length = len(str(val)) if val is not None else 0
                    if length > max_length:
                        max_length = length
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    output.seek(0)
    return output.read()


def create_pdf_bytes(results: List[dict], totals: Tuple[float, float, float, float]) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        raise RuntimeError("reportlab não instalado. Execute: pip install reportlab")

    header = ["Conta", "Excel Prev", "PDF Prev", "Dif Prev", "Status Prev",
              "Excel Curr", "PDF Curr", "Dif Curr", "Status Curr", "Arquivo PDF"]
    data = [header]
    style_list = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d3d3d3')),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER')
    ]

    for r in results:
        row = [
            r['account'],
            format_brazilian(r['excel_prev']),
            format_brazilian(r['pdf_prev']),
            format_brazilian(r['diff_prev']),
            r['status_prev'],
            format_brazilian(r['excel_curr']),
            format_brazilian(r['pdf_curr']),
            format_brazilian(r['diff_curr']),
            r['status_curr'],
            r['pdf_file']
        ]
        data.append(row)
        if r['status_prev'] == 'NÃO CONFERE' or r['status_curr'] == 'NÃO CONFERE':
            style_list.append(('BACKGROUND', (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor('#fce4e4')))

    total_prev_excel, total_prev_pdf, total_curr_excel, total_curr_pdf = totals
    totals_row = [
        "Totais",
        format_brazilian(total_prev_excel),
        format_brazilian(total_prev_pdf),
        "",
        "",
        format_brazilian(total_curr_excel),
        format_brazilian(total_curr_pdf),
        "",
        "",
        ""
    ]
    data.append(totals_row)
    totals_index = len(data) - 1
    style_list.append(('BACKGROUND', (0, totals_index), (-1, totals_index), colors.HexColor('#EDEDED')))
    style_list.append(('FONTNAME', (0, totals_index), (-1, totals_index), 'Helvetica-Bold'))

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = [Paragraph("Relatório de Confronto de Saldos", styles['Title']), Spacer(1, 12)]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle(style_list))
    story.append(table)
    doc.build(story)
    output.seek(0)
    return output.read()


# -----------------------
# Streamlit UI
# -----------------------
st.set_page_config(page_title="Confronto de Saldos", layout="wide")
st.title("Confronto de Saldos — v17 (Web com OCR & AgGrid)")

# session state initial
for key, default in [
    ('progress_percent', 0), ('processed_count', 0),
    ('cancel_requested', False), ('last_results', []),
    ('last_totals', (0.0, 0.0, 0.0, 0.0))
]:
    if key not in st.session_state:
        st.session_state[key] = default

# Sidebar inputs
with st.sidebar:
    st.header("Entradas")
    uploaded_excel = st.file_uploader("Upload do Excel (Conta | Saldo Anterior | Saldo Atual)", type=['xlsx', 'xls'])
    uploaded_pdfs = st.file_uploader("Upload dos PDFs (múltiplos)", type=['pdf'], accept_multiple_files=True)
    st.markdown("---")
    st.write("Opções de OCR/Processamento:")
    col1, col2 = st.columns(2)
    with col1:
        use_ocr = st.checkbox("Ativar OCR automático (fallback)", value=True)
    with col2:
        st.write("")  # placeholder
    st.markdown("---")
    st.markdown("Instruções:\n- Planilha: 1ª coluna Conta, 2ª Saldo Anterior, 3ª Saldo Atual\n- Faça upload do Excel e ao menos 1 PDF")

# Load Excel
df_accounts = None
if uploaded_excel is not None:
    try:
        df_tmp = pd.read_excel(uploaded_excel, header=0)
        if df_tmp.shape[1] < 3:
            st.sidebar.error("A planilha precisa ter ao menos 3 colunas.")
        else:
            df2 = df_tmp.iloc[:, :3].copy()
            df2.columns = ['account', 'excel_prev', 'excel_curr']
            def to_float(v):
                if pd.isna(v):
                    return None
                if isinstance(v, (int, float)):
                    return float(v)
                s = str(v).strip().replace('.', '').replace(',', '.')
                try:
                    return float(s)
                except:
                    return None
            df2['excel_prev'] = df2['excel_prev'].apply(to_float)
            df2['excel_curr'] = df2['excel_curr'].apply(to_float)
            df2['account'] = df2['account'].astype(str).str.strip()
            df_accounts = df2
            st.sidebar.success(f"Excel carregado: {len(df2)} contas.")
    except Exception as e:
        st.sidebar.error(f"Falha ao ler Excel: {e}")

# Load PDF bytes
pdf_bytes_list = []
if uploaded_pdfs:
    for up in uploaded_pdfs:
        try:
            b = up.read()
            pdf_bytes_list.append(b)
        except Exception:
            pass
    if pdf_bytes_list:
        st.sidebar.success(f"{len(pdf_bytes_list)} PDF(s) carregado(s).")

# Main area: controls/progress/results
col1, col2 = st.columns([1, 1])
with col1:
    st.subheader("Ações")
    run_button = st.button("Executar Confronto")
    cancel_button = st.button("Cancelar")
    if cancel_button:
        st.session_state['cancel_requested'] = True
        st.success("Cancelamento solicitado.")

with col2:
    st.subheader("Progresso")
    percent = st.session_state.get('progress_percent', 0)
    processed_count = st.session_state.get('processed_count', 0)
    st.progress(percent)
    st.write(f"{percent}% — {processed_count} processadas")

# Run processing (synchronously but cooperative - Streamlit single-threaded per session)
if run_button:
    st.session_state['cancel_requested'] = False
    st.session_state['progress_percent'] = 0
    st.session_state['processed_count'] = 0
    st.session_state['last_results'] = []
    st.session_state['last_totals'] = (0.0, 0.0, 0.0, 0.0)

    if df_accounts is None:
        st.error("Carregue o Excel antes de executar.")
    elif len(pdf_bytes_list) == 0:
        st.error("Carregue ao menos 1 PDF antes de executar.")
    else:
        # process
        with st.spinner("Processando..."):
            # if OCR disabled, temporarily monkey-patch to only use pdfplumber
            if not use_ocr:
                # we'll try pdfplumber only: set convert_from_bytes/pytesseract to None fallback
                saved_convert = globals().get('convert_from_bytes', None)
                saved_pyt = globals().get('pytesseract', None)
                # disable OCR by setting them None in the module scope for extract_text function
                globals()['convert_from_bytes'] = None
                globals()['pytesseract'] = None

                results, totals, cancelled = process_confronto_streamlit(df_accounts, pdf_bytes_list)

                # restore
                globals()['convert_from_bytes'] = saved_convert
                globals()['pytesseract'] = saved_pyt
            else:
                results, totals, cancelled = process_confronto_streamlit(df_accounts, pdf_bytes_list)

        st.session_state['last_results'] = results
        st.session_state['last_totals'] = totals

        if cancelled:
            st.warning("Processamento cancelado. Resultados parciais exibidos.")
        else:
            st.success("Confronto concluído.")

# Display results using AgGrid if available, else fallback to st.dataframe
results_display = st.session_state.get('last_results', [])
totals = st.session_state.get('last_totals', (0.0, 0.0, 0.0, 0.0))

st.markdown("---")
st.subheader("Resultados")

if results_display:
    df_display_raw = pd.DataFrame(results_display)
    
    # <-- CORREÇÃO APLICADA AQUI
    # Converte 'np.nan' (não serializável) para 'None' (serializável)
    # Isso corrige o 'MarshallComponentException' no AgGrid
    df_display = df_display_raw.astype(object).where(pd.notnull(df_display_raw), None)

    # Cria cópia para exibição formatada
    df_display_formatted = pd.DataFrame()
    
    # format columns for display
    df_display_formatted['Excel Prev'] = df_display['excel_prev'].apply(format_brazilian)
    df_display_formatted['PDF Prev'] = df_display['pdf_prev'].apply(format_brazilian)
    df_display_formatted['Dif Prev'] = df_display['diff_prev'].apply(format_brazilian)
    df_display_formatted['Excel Curr'] = df_display['excel_curr'].apply(format_brazilian)
    df_display_formatted['PDF Curr'] = df_display['pdf_curr'].apply(format_brazilian)
    df_display_formatted['Dif Curr'] = df_display['diff_curr'].apply(format_brazilian)
    df_display_formatted['Status Prev'] = df_display['status_prev']
    df_display_formatted['Status Curr'] = df_display['status_curr']
    df_display_formatted['Conta'] = df_display['account']
    
    df_display_formatted = df_display_formatted[[
        'Conta', 'Excel Prev', 'PDF Prev', 'Dif Prev', 'Status Prev',
        'Excel Curr', 'PDF Curr', 'Dif Curr', 'Status Curr'
    ]]

    if AgGrid is not None and GridOptionsBuilder is not None:
        gb = GridOptionsBuilder.from_dataframe(df_display_formatted) # Passar o DF formatado
        gb.configure_default_column(aggregate=True, groupable=False, value=True, enableRowGroup=False, resizable=True)
        # center-align
        gb.configure_column("Conta", header_name="Conta", cellStyle={'textAlign': 'center'})
        # quick cell style based on status columns (JsCode)
        js_cell_style = JsCode("""
        function(params) {
            if (params.data['Status Prev'] && params.data['Status Prev'].includes('NÃO CONFERE')) {
                return {'backgroundColor':'#f7d4d4'}
            }
            if (params.data['Status Curr'] && params.data['Status Curr'].includes('NÃO CONFERE')) {
                return {'backgroundColor':'#f7d4d4'}
            }
            if (params.data['Status Prev'] && params.data['Status Prev'].includes('CONFERE')) {
                return {'backgroundColor':'#d4f7d4'}
            }
            if (params.data['Status Curr'] && params.data['Status Curr'].includes('CONFERE')) {
                return {'backgroundColor':'#d4f7d4'}
            }
            return null;
        }
        """)
        gb.configure_default_column(cellStyle=js_cell_style)
        gridOptions = gb.build()
        # Passar o DF formatado para o AgGrid
        AgGrid(df_display_formatted, gridOptions=gridOptions, height=400, fit_columns_on_grid_load=True, update_mode=GridUpdateMode.NO_UPDATE)
    else:
        # fallback plain
        st.dataframe(df_display_formatted, use_container_width=True)

    # totals
    st.markdown("**Totais (por fonte)**")
    c1, c2 = st.columns(2)
    with c1:
        st.write(f"Saldo Anterior (Excel): **{format_brazilian(totals[0])}**")
        st.write(f"Saldo Atual (Excel): **{format_brazilian(totals[2])}**")
    with c2:
        st.write(f"Saldo Anterior (PDF): **{format_brazilian(totals[1])}**")
        st.write(f"Saldo Atual (PDF): **{format_brazilian(totals[3])}**")

    # export buttons
    st.markdown("---")
    st.subheader("Exportar Relatório")
    colx, coly = st.columns(2)
    with colx:
        try:
            # Passar os resultados originais (lista de dicts) para a função de export
            excel_bytes = create_excel_bytes(results_display, totals)
            st.download_button("Exportar Excel (.xlsx)", data=excel_bytes, file_name="Relatorio_Confronto.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erro ao gerar Excel: {e}")
    with coly:
        try:
            # Passar os resultados originais (lista de dicts) para a função de export
            pdf_bytes = create_pdf_bytes(results_display, totals)
            st.download_button("Exportar PDF (.pdf)", data=pdf_bytes, file_name="Relatorio_Confronto.pdf", mime="application/pdf")
        except Exception as e:
            st.error(f"Erro ao gerar PDF: {e}")
else:
    st.info("Nenhum resultado ainda. Faça upload do Excel e PDFs e clique em 'Executar Confronto'.")

st.markdown("---")
st.caption("Para PDFs escaneados, o app tenta usar OCR automaticamente (se pytesseract + pdf2image estiverem disponíveis).")
